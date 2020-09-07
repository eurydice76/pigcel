"""This module implements the following classes and functions:
    - InvalidExcelWorkbookError
    - UnknownPropertyError
    - InvalidTimeError
    - ExcelWorkbookReader
"""

import collections
import os
import sys

import openpyxl

import pandas as pd

import numpy as np


class InvalidExcelWorkbookError(Exception):
    """This class implements an exception raised when the workbook is invalid.
    """


class UnknownPropertyError(Exception):
    """This class implements an exception raised when a requested property is unknown.
    """


class InvalidTimeError(Exception):
    """This class implements an exception raised when a requested time is not valid.
    """


class ExcelWorkbookReader:
    """This class implements the reader for the data stored in excel file. To be valid the excel file must contain respectively
    'Suivi', 'Data', 'Gaz du sang' and 'NFS' sheets.
    """

    times = ['T-30', 'T0', 'T10', 'T20', 'T30', 'T1H', 'T1H30', 'T2H', 'T3H', 'T4H', 'T5H', 'T6H']

    def __init__(self, filename):
        """Constructor

        Args:
            filename (str): the path to the workbook
        """

        self._workbook = openpyxl.load_workbook(filename)

        sheet_names = set(self._workbook.get_sheet_names())

        if len(sheet_names.intersection(['Suivi', 'Data', 'Gaz du sang', 'NFS'])) != 4:
            raise InvalidExcelWorkbookError('One or more compulsory sheets are missing.')

        self._filename = filename

    @property
    def properties(self):
        """Returns the properties stored in the workbook. Only the sheet 'Data', 'Gaz du sang' and 'NFS' are considered.

        Returns:
            list of str: the properties
        """

        properties = []

        data_sheet = self._workbook.get_sheet_by_name('Data')
        properties.extend([data_sheet.cell(row=6, column=col).value for col in range(3, 22)])

        blood_gas_sheet = self._workbook.get_sheet_by_name('Gaz du sang')
        properties.extend([blood_gas_sheet.cell(row=row, column=1).value for row in range(6, 31)])

        nfs_sheet = self._workbook.get_sheet_by_name('NFS')
        properties.extend([nfs_sheet.cell(row=row, column=3).value for row in range(3, 17)])

        return properties

    @property
    def basename(self):

        return os.path.splitext(os.path.basename(self._filename))[0]

    @property
    def filename(self):
        """Returns the workbook's filename.

        Returns:
            str: the filename
        """

        return self._filename

    def get_data(self):
        """Returns the data stored in the workbook. Only the sheet 'Data', 'Gaz du sang' and 'NFS' are considered.

        Returns:
            pandas.DataFrame: the data
        """

        data = self.parse_data_worksheet()
        data = pd.concat([data, self.parse_blood_gas_worksheet()], axis=1)
        data = pd.concat([data, self.parse_nfs_worksheet()], axis=1)

        return data

    def get_property_slice(self, selected_property, selected_times=None):
        """Return a slice of the data according to a given property and selected times.

        Args:
            selected_property (str): the selected property
            times (list of str): the selected times

        Returns:
            pd.Series: the slice
        """

        data = self.get_data()

        if selected_property not in data.columns:
            raise UnknownPropertyError('The property {} is unknown'.format(selected_property))

        if selected_times is None:
            selected_times = data.index
        else:
            temp = []
            for s in selected_times:
                if s not in data.index:
                    continue
                temp.append(s)
            selected_times = pd.Index(temp)

        if selected_times.empty:
            raise InvalidTimeError('The selected times are not valid times')

        data = data[selected_property].loc[selected_times]

        return data

    def get_time_slice(self, selected_time):
        """Return a slice of the data according to a given time.

        Args:
            selected_time (str): the time

        Returns:
            pd.Series: the slice
        """

        if selected_time not in ExcelWorkbookReader.times:
            raise InvalidTimeError('The time {} is not a registered time'.format(selected_time))

        data = self.get_data()
        data = data.loc[selected_time]

        return data

    @ property
    def information(self):
        """Parse the sheet 'Suivi' and stringify the data stored in it.

        Returns:
            str: the general information about the animal
        """

        data_sheet = self._workbook.get_sheet_by_name('Suivi')

        cells = data_sheet['A1':'B13']

        info = []
        for cell1, cell2 in cells:
            info.append(': '.join([str(cell1.value), str(cell2.value)]))

        info = '\n'.join(info)

        return info

    def parse_blood_gas_worksheet(self):
        """Parse the sheet 'Gaz du sang' and fetch the data stored in it.

        Returns:
            collections.OrderedDict: the data
        """

        data_sheet = self._workbook.get_sheet_by_name('Gaz du sang')

        properties = [data_sheet.cell(row=row, column=1).value for row in range(6, 31)]

        cells = data_sheet['B6':'M30']

        cells = list(zip(*cells))

        data = pd.DataFrame([[cell.value if cell.value is not None else np.nan for cell in row]
                             for row in cells], columns=properties, index=ExcelWorkbookReader.times)

        return data

    def parse_data_worksheet(self):
        """Parse the sheet 'Data' and fetch the data stored in it.

        Returns:
            collections.OrderedDict: the data
        """

        data_sheet = self._workbook.get_sheet_by_name('Data')

        properties = [data_sheet.cell(row=6, column=col).value for col in range(3, 22)]

        cells = data_sheet['C7':'U18']

        data = pd.DataFrame([[cell.value if cell.value is not None else np.nan for cell in row]
                             for row in cells], columns=properties, index=ExcelWorkbookReader.times)

        return data

    def parse_nfs_worksheet(self):
        """Parse the sheet 'NFS' and fetch the data stored in it.

        Returns:
            collections.OrderedDict: the data
        """

        data_sheet = self._workbook.get_sheet_by_name('NFS')

        properties = [data_sheet.cell(row=row, column=3).value for row in range(3, 17)]

        cells = data_sheet['D3':'M16']

        data = collections.OrderedDict()

        selected_times = [0, 1, 5, 8, 11]
        selected_values = [0, 2, 4, 6, 8]

        data = []
        for row, _ in enumerate(properties):

            values = [cell.value for cell in cells[row]]

            row = [np.nan]*len(self.times)

            for i in range(len(selected_values)):
                row[selected_times[i]] = values[selected_values[i]]

            data.append(row)

        data = pd.DataFrame(data, index=properties, columns=ExcelWorkbookReader.times).T
        data.columns = properties
        data.set_index(pd.Series(ExcelWorkbookReader.times), inplace=True)

        return data


if __name__ == '__main__':

    workbook = sys.argv[1]

    mwb = ExcelWorkbookReader(workbook)

    print(mwb.information)

    print(mwb.parse_data_worksheet())

    print(mwb.parse_blood_gas_worksheet())

    print(mwb.parse_nfs_worksheet())

    print(mwb.get_data())

    print(mwb.get_property_slice('FC'))

    print(mwb.get_time_slice('T0'))
