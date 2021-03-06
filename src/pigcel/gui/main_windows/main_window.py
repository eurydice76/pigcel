"""This modules implements the following classes:
    _ MainWindow
"""

import collections
import glob
import logging
import os
import sys

from PyQt5 import QtCore, QtGui, QtWidgets

import numpy as np

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

import pigcel
from pigcel.__pkginfo__ import __version__
from pigcel.gui.models.animals_data_model import AnimalsDataModel
from pigcel.gui.models.workbook_data_model import WorkbookDataModel
from pigcel.gui.models.animals_groups_model import AnimalsGroupsModel
from pigcel.gui.views.animals_data_listview import AnimalsDataListView
from pigcel.gui.views.copy_pastable_tableview import CopyPastableTableView
from pigcel.gui.widgets.logger_widget import QTextEditLogger
from pigcel.gui.widgets.groups_widget import GroupsWidget
from pigcel.gui.widgets.multiple_directories_selector import MultipleDirectoriesSelector
from pigcel.gui.widgets.plots_widget import PlotsWidget
from pigcel.kernel.readers.excel_reader import ExcelWorkbookReader
from pigcel.kernel.utils.progress_bar import progress_bar


class MainWindow(QtWidgets.QMainWindow):
    """This class implements the main window of the application.
    """

    add_new_group = QtCore.pyqtSignal(str)

    display_group_averages = QtCore.pyqtSignal()

    display_group_effect_statistics = QtCore.pyqtSignal()

    display_time_effect_statistics = QtCore.pyqtSignal()

    display_group_medians = QtCore.pyqtSignal()

    export_group_statistics = QtCore.pyqtSignal(str,str)

    import_groups_from_directories = QtCore.pyqtSignal(dict)

    update_property_plot = QtCore.pyqtSignal(tuple)

    update_time_plot = QtCore.pyqtSignal(tuple)

    def __init__(self, parent=None):
        """Constructor.

        Args:
            parent (QtCore.QObject): the parent window
        """

        super(MainWindow, self).__init__(parent)

        self.init_ui()

    def build_events(self):
        """Build the signal/slots.
        """

        self._animals_list.double_clicked_empty.connect(self.on_load_monitoring_file)
        self._selected_property_combo.currentTextChanged.connect(self.on_update_property_plot)
        self._selected_property_combo.currentTextChanged.connect(self.on_update_time_plot)
        self._selected_time_combo.currentTextChanged.connect(self.on_update_time_plot)

    def build_layout(self):
        """Build the layout.
        """

        main_layout = QtWidgets.QVBoxLayout()

        hlayout = QtWidgets.QHBoxLayout()

        vlayout = QtWidgets.QVBoxLayout()
        vlayout.addWidget(self._animals_list)

        property_layout = QtWidgets.QHBoxLayout()
        property_layout.addWidget(self._selected_property_label)
        property_layout.addWidget(self._selected_property_combo)
        vlayout.addLayout(property_layout)

        time_layout = QtWidgets.QHBoxLayout()
        time_layout.addWidget(self._selected_time_label)
        time_layout.addWidget(self._selected_time_combo)
        vlayout.addLayout(time_layout)

        hlayout.addLayout(vlayout, stretch=1)

        hlayout.addWidget(self._tabs, stretch=4)

        main_layout.addLayout(hlayout, stretch=4)

        main_layout.addWidget(self._data_table, stretch=2)

        main_layout.addWidget(self._logger.widget, stretch=1)

        self._main_frame.setLayout(main_layout)

    def build_menu(self):
        """Build the menu.
        """

        menubar = self.menuBar()

        file_menu = menubar.addMenu('&File')

        file_action = QtWidgets.QAction('&Open monitoring file', self)
        file_action.setShortcut('Ctrl+O')
        file_action.setStatusTip('Open monitoring (xlsx) file')
        file_action.triggered.connect(self.on_load_monitoring_file)
        file_menu.addAction(file_action)

        file_menu.addSeparator()

        exit_action = QtWidgets.QAction('&Exit', self)
        exit_action.setShortcut('Ctrl+Q')
        exit_action.setStatusTip('Exit pigcel')
        exit_action.triggered.connect(self.on_quit_application)
        file_menu.addAction(exit_action)

        group_menu = menubar.addMenu('&Groups')
        add_group_action = QtWidgets.QAction('&Add group', self)
        add_group_action.setShortcut('Ctrl+R')
        add_group_action.setStatusTip('Add new group')
        add_group_action.triggered.connect(self.on_add_new_group)
        group_menu.addAction(add_group_action)

        import_groups_action = QtWidgets.QAction('&Import from directories', self)
        import_groups_action.setShortcut('Ctrl+U')
        import_groups_action.setStatusTip('Import and create groups from a list of directories')
        import_groups_action.triggered.connect(self.on_import_groups_from_directories)
        group_menu.addAction(import_groups_action)

        group_menu.addSeparator()

        display_group_averages_action = QtWidgets.QAction('&Display averages', self)
        display_group_averages_action.setShortcut('Ctrl+D')
        display_group_averages_action.setStatusTip('Display averages and std for each group')
        display_group_averages_action.triggered.connect(self.on_display_group_averages)
        group_menu.addAction(display_group_averages_action)

        display_group_medians_action = QtWidgets.QAction('Display &medians', self)
        display_group_medians_action.setShortcut('Ctrl+M')
        display_group_medians_action.setStatusTip('Display averages and std for each group')
        display_group_medians_action.triggered.connect(self.on_display_group_medians)
        group_menu.addAction(display_group_medians_action)

        export_group_statistics = QtWidgets.QAction('&Export descriptive statistics', self)
        export_group_statistics.setShortcut('Ctrl+E')
        export_group_statistics.setStatusTip('Export descriptive statistics (average, std, quartile ...)')
        export_group_statistics.triggered.connect(self.on_export_group_statistics)
        group_menu.addAction(export_group_statistics)

        export_all_group_statistics = QtWidgets.QAction('&Export all descriptive statistics', self)
        export_all_group_statistics.setShortcut('Ctrl+A')
        export_all_group_statistics.setStatusTip('Export descriptive statistics (average, std, quartile ...)')
        export_all_group_statistics.triggered.connect(self.on_export_all_group_statistics)
        group_menu.addAction(export_all_group_statistics)

        statistics_menu = menubar.addMenu('&Statistics')

        group_effect_menu = statistics_menu.addMenu('&Groups effect')

        group_effect_action = QtWidgets.QAction('&Compute', self)
        group_effect_action.setShortcut('Ctrl+G')
        group_effect_action.setStatusTip('Display group effect statistics')
        group_effect_action.triggered.connect(self.on_display_group_effect_statistics)
        group_effect_menu.addAction(group_effect_action)

        export_all_group_effects_action = QtWidgets.QAction('&Export all', self)
        export_all_group_effects_action.setShortcut('Ctrl+B')
        export_all_group_effects_action.setStatusTip('Export group effects for all properties')
        export_all_group_effects_action.triggered.connect(self.on_export_all_group_effects)
        group_effect_menu.addAction(export_all_group_effects_action)

        time_effect_menu = statistics_menu.addMenu('&Time effect')

        time_effect_action = QtWidgets.QAction('&Time effect', self)
        time_effect_action.setShortcut('Ctrl+T')
        time_effect_action.setStatusTip('Display time effect statistics')
        time_effect_action.triggered.connect(self.on_display_time_effect_statistics)
        time_effect_menu.addAction(time_effect_action)

        export_all_time_effects_action = QtWidgets.QAction('&Export all', self)
        export_all_time_effects_action.setShortcut('Ctrl+L')
        export_all_time_effects_action.setStatusTip('Export time effects for all properties')
        export_all_time_effects_action.triggered.connect(self.on_export_all_time_effects)
        time_effect_menu.addAction(export_all_time_effects_action)


    def build_widgets(self):
        """Build the widgets.
        """

        self._main_frame = QtWidgets.QFrame(self)

        self._animals_list = AnimalsDataListView()
        self._animals_list.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self._animals_list.setDragEnabled(True)
        self._animals_list.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self._animals_list.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)

        animals_model = AnimalsDataModel(self)
        self._animals_list.setModel(animals_model)

        self._selected_property_label = QtWidgets.QLabel('Property')
        self._selected_property_combo = QtWidgets.QComboBox()

        self._selected_time_label = QtWidgets.QLabel('Time')
        self._selected_time_combo = QtWidgets.QComboBox()

        self._data_table = CopyPastableTableView()
        self._data_table.setSelectionBehavior(QtWidgets.QTableView.SelectRows)
        self._data_table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)

        self._logger = QTextEditLogger(self)
        self._logger.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        logging.getLogger().addHandler(self._logger)
        logging.getLogger().setLevel(logging.INFO)

        self._tabs = QtWidgets.QTabWidget()

        self._plots_widget = PlotsWidget(self)
        self._groups_widget = GroupsWidget(animals_model, self)

        self._tabs.addTab(self._plots_widget, 'Plots')
        self._tabs.addTab(self._groups_widget, 'Groups')

        self.setCentralWidget(self._main_frame)

        self.setGeometry(0, 0, 1200, 1100)

        self.setWindowTitle('pigcel {}'.format(__version__))

        self._progress_label = QtWidgets.QLabel('Progress')
        self._progress_bar = QtWidgets.QProgressBar()
        progress_bar.set_progress_widget(self._progress_bar)
        self.statusBar().showMessage('pigcel {}'.format(__version__))
        self.statusBar().addPermanentWidget(self._progress_label)
        self.statusBar().addPermanentWidget(self._progress_bar)

        icon_path = os.path.join(pigcel.__path__[0], "icons", "pigcel.png")
        self.setWindowIcon(QtGui.QIcon(icon_path))

        self.show()

    def init_ui(self):
        """Initializes the ui.
        """

        self.build_widgets()

        self.build_layout()

        self.build_menu()

        self.build_events()

    def on_add_new_group(self):
        """Event fired when the user clicks on 'Add group' menu button.
        """

        group, ok = QtWidgets.QInputDialog.getText(self, 'Enter group name', 'Group name', QtWidgets.QLineEdit.Normal, 'group')

        if ok and group:
            self.add_new_group.emit(group)

    def on_display_group_averages(self):
        """Event fired when the user clicks on 'Display group averages plot' menu button.
        """

        self.display_group_averages.emit()

    def on_display_group_effect_statistics(self):
        """Event fire when the user clicks on 'Display group effect' menu button.
        """

        self.display_group_effect_statistics.emit()

    def on_display_time_effect_statistics(self):
        """Event fire when the user clicks on 'Display time effect' menu button.
        """

        self.display_time_effect_statistics.emit()

    def on_display_group_medians(self):
        """Event fired when the user clicks on 'Display group medians plot' menu button.
        """

        self.display_group_medians.emit()

    def on_export_all_group_effects(self):
        """Export the group effect computed for all properties.
        """

        all_properties = [self._selected_property_combo.itemText(index) for index in range(self._selected_property_combo.count())]
        if not all_properties:
            logging.info('No properties loaded')
            return

        groups_model = self._groups_widget.model()
        if groups_model is None:
            logging.info('No groups defined')
            return

        filename, _ = QtWidgets.QFileDialog.getSaveFileName(self, caption='Export statistics as ...', filter="Excel files (*.xls *.xlsx)")
        if not filename:
            return

        filename_noext, ext = os.path.splitext(filename)
        if ext not in ['.xls', '.xlsx']:
            logging.warning('Bad file extension for output excel file {}. It will be replaced by ".xlsx"'.format(filename))
            filename = filename_noext + '.xlsx'

        workbook = openpyxl.Workbook()
        # Remove the first empty sheet created by default
        workbook.remove_sheet(workbook.get_sheet_by_name('Sheet'))

        progress_bar.reset(len(all_properties))

        n_groups = groups_model.rowCount()

        for i, selected_property in enumerate(all_properties):

            # Create the excel worksheet for the selected property
            worksheet = workbook.create_sheet(selected_property)

            global_effect = groups_model.evaluate_global_group_effect(selected_property)

            for j, col in enumerate(global_effect.columns):
                worksheet.cell(1,j+2).value = col

            for j, row in enumerate(global_effect.index):
                worksheet.cell(j+2,1).value = row

            for j in range(len(global_effect.index)):
                for k in range(len(global_effect.columns)):
                    val = 'nan' if np.isnan(global_effect.iloc[j,k]).any() else global_effect.iloc[j,k]
                    worksheet.cell(j+2,k+2).value = val

            current_row = global_effect.shape[0] + 3

            pairwise_group_effect = groups_model.evaluate_pairwise_group_effect(selected_property)

            for j, (time, pairwise_effect) in enumerate(pairwise_group_effect.items()):
                row = current_row + j*(pairwise_effect.shape[0] + 3)
                worksheet.cell(row,1).value = 't = {}'.format(time)
                for k, col_name in enumerate(pairwise_effect.columns):
                    worksheet.cell(row+1,k+2).value = col_name
                for k, row_name in enumerate(pairwise_effect.index):
                    worksheet.cell(row+k+2,1).value = row_name
                
                for k in range(len(pairwise_effect.index)):
                    for l in range(len(pairwise_effect.columns)):
                        val = 'nan' if np.isnan(pairwise_effect.iloc[k,l]).any() else pairwise_effect.iloc[k,l]
                        worksheet.cell(row+k+2,l+2).value = val

            progress_bar.update(i+1)

        try:
            workbook.save(filename)
        except PermissionError as error:
            logging.error('Can not save {} notebook: {}'.format(filename, str(error)))
            return

    def on_export_all_time_effects(self):
        """Export the group effect computed for all properties.
        """

        all_properties = [self._selected_property_combo.itemText(index) for index in range(self._selected_property_combo.count())]
        if not all_properties:
            logging.info('No properties loaded')
            return

        groups_model = self._groups_widget.model()
        if groups_model is None:
            logging.info('No groups defined')
            return

        filename, _ = QtWidgets.QFileDialog.getSaveFileName(self, caption='Export statistics as ...', filter="Excel files (*.xls *.xlsx)")
        if not filename:
            return

        filename_noext, ext = os.path.splitext(filename)
        if ext not in ['.xls', '.xlsx']:
            logging.warning('Bad file extension for output excel file {}. It will be replaced by ".xlsx"'.format(filename))
            filename = filename_noext + '.xlsx'

        workbook = openpyxl.Workbook()
        # Remove the first empty sheet created by default
        workbook.remove_sheet(workbook.get_sheet_by_name('Sheet'))

        progress_bar.reset(len(all_properties))

        # n_groups = groups_model.rowCount()

        for i, selected_property in enumerate(all_properties):

            # Create the excel worksheet for the selected property
            worksheet = workbook.create_sheet(selected_property)

            # Compute the global time effect for the running property
            times_per_group, global_effect = groups_model.evaluate_global_time_effect(selected_property)

            # Compute the pairwise time effect for the running property
            pairwise_effect = groups_model.evaluate_pairwise_time_effect(selected_property)

            # Loop over the groups and write the data per block
            # A block is made of:
            #   1st line: the name of the group
            #   2nd line: the list of valid times (time for which the pairwise effect could be computed)
            #   3rd line: the p value for the global effect
            # Then the p-value 12x12 matrix for the pairwise effect (all times included from -0h30 to 6h00)
            comp = 0
            for j, (group,pairwise_effect) in enumerate(pairwise_effect.items()):

                comp += 1

                worksheet.cell(comp,1).value = 'Group'
                worksheet.cell(comp,2).value = group

                comp += 1

                # These are the times for which the time effect could be computed
                valid_times = [time[0] for time in times_per_group[group] if time[1]]

                worksheet.cell(comp,1).value = 'Valid times'
                for k, valid_time in enumerate(valid_times):
                    worksheet.cell(comp,k+2).value = valid_time

                comp += 1
                worksheet.cell(comp,1).value = 'p-value'
                worksheet.cell(comp,2).value = global_effect.loc[group,'p-value']

                comp += 2

                for k, time in enumerate(pairwise_effect.columns):
                    worksheet.cell(comp,k+2).value = time

                for k, time in enumerate(pairwise_effect.index):
                    comp += 1
                    worksheet.cell(comp,1).value = time
                    for l in range(len(pairwise_effect.columns)):
                        val = 'nan' if np.isnan(pairwise_effect.iloc[k,l]).any() else pairwise_effect.iloc[k,l]
                        worksheet.cell(comp,l+2).value = val

                comp += 2

            progress_bar.update(i+1)

        try:
            workbook.save(filename)
        except PermissionError as error:
            logging.error('Can not save {} notebook: {}'.format(filename, str(error)))
            return

    def _export_descriptive_statistics(self, selected_property, filename):
        """
        """

        self.export_group_statistics.emit(selected_property, filename)

    def on_export_all_group_statistics(self):
        """Event fired when the user clicks on the 'Export all descriptive statistics' menu button
        """

        selected_directory = QtWidgets.QFileDialog.getExistingDirectory(self, "Select Directory")
        if not selected_directory:
            return

        # Check that the selected directory is writable
        try:
            test_file = os.path.join(selected_directory,'xxx')
            fout = open(test_file,'w')
        except IOError as e:
            logging.error(str(e))
            return
        else:
            fout.close()
            os.remove(test_file)

        all_properties = [self._selected_property_combo.itemText(i) for i in range(self._selected_property_combo.count())]

        progress_bar.reset(len(all_properties))

        for i, prop in enumerate(all_properties):
            filename = os.path.join(selected_directory,'{}.xlsx'.format(prop))
            self._export_descriptive_statistics(prop, filename)
            progress_bar.update(i+1)

    def on_export_group_statistics(self):
        """Event fired when the user clicks on the 'Export statistics' menu button.
        """

        filename, _ = QtWidgets.QFileDialog.getSaveFileName(self, caption='Export statistics as ...', filter="Excel files (*.xls *.xlsx)")
        if not filename:
            return

        filename_noext, ext = os.path.splitext(filename)
        if ext not in ['.xls', '.xlsx']:
            logging.warning('Bad file extension for output excel file {}. It will be replaced by ".xlsx"'.format(filename))
            filename = filename_noext + '.xlsx'

        self._export_descriptive_statistics(self.selected_property, filename)

    def on_import_groups_from_directories(self):
        """Event fired when the user clicks on Groups -> Import from directories menu button.
        """

        # Pop up a file browser
        selector = MultipleDirectoriesSelector()
        if not selector.exec_():
            return

        experimental_dirs = selector.selectedFiles()
        if not experimental_dirs:
            return

        animals_data_model = self._animals_list.model()

        progress_bar.reset(len(experimental_dirs))

        n_loaded_dirs = 0

        groups = collections.OrderedDict()

        all_properties = set([self._selected_property_combo.itemText(index) for index in range(self._selected_property_combo.count())])
        all_times = set([self._selected_time_combo.itemText(index) for index in range(self._selected_time_combo.count())])

        # Loop over the pig directories
        for progress, exp_dir in enumerate(experimental_dirs):

            data_files = glob.glob(os.path.join(exp_dir, '*.xls[x]'))

            # Loop over the Data*csv csv files found in the current directory
            for data_file in data_files:
                try:
                    reader = ExcelWorkbookReader(data_file)
                except Exception as error:
                    logging.error('Error reading {} file: {}'.format(data_file, str(error)))
                    continue
                else:
                    animals_data_model.add_workbook(reader)
                    groups.setdefault(os.path.basename(exp_dir), []).append(data_file)
                    all_properties.update(reader.data.columns)
                    all_times.update(reader.data.index)

            n_loaded_dirs += 1
            progress_bar.update(progress+1)

        self._selected_property_combo.clear()
        self._selected_property_combo.addItems(sorted(all_properties))

        self._selected_time_combo.clear()
        self._selected_time_combo.addItems(sorted(all_times))

        # Create a signal/slot connexion for row changed event
        self._animals_list.selectionModel().selectionChanged.connect(self.on_select_animal)

        self._animals_list.setCurrentIndex(animals_data_model.index(0, 0))

        self.import_groups_from_directories.emit(groups)

        logging.info('Imported successfully {} groups out of {} directories'.format(n_loaded_dirs, len(experimental_dirs)))

    def on_load_monitoring_file(self):
        """Event called when the user clicks on 'Load monitoring file' menu button.
        """

        # Pop up a file browser for selecting the workbooks
        xlsx_files = QtWidgets.QFileDialog.getOpenFileNames(self, 'Open data files', '', 'Data Files (*.xls *.xlsx)')[0]
        if not xlsx_files:
            return

        workbooks_model = self._animals_list.model()

        n_xlsx_files = len(xlsx_files)
        progress_bar.reset(n_xlsx_files)

        n_loaded_files = 0

        all_properties = set([self._selected_property_combo.itemText(index) for index in range(self._selected_property_combo.count())])
        all_times = set([self._selected_time_combo.itemText(index) for index in range(self._selected_time_combo.count())])

        # Loop over the pig directories
        for progress, xlsx_file in enumerate(xlsx_files):

            # Any error at reading must be caught here
            try:
                reader = ExcelWorkbookReader(xlsx_file)
            except Exception as error:
                logging.error('Error reading {} file: {}'.format(xlsx_file, str(error)))
                continue
            else:
                workbooks_model.add_workbook(reader)
                n_loaded_files += 1
            finally:
                self._selected_property_combo.clear()
                progress_bar.update(progress+1)

        self._selected_property_combo.clear()
        self._selected_property_combo.addItems(sorted(all_properties))

        self._selected_time_combo.clear()
        self._selected_time_combo.addItems(sorted(all_times))

        # Create a signal/slot connexion for row changed event
        self._animals_list.selectionModel().selectionChanged.connect(self.on_select_animal)

        self._animals_list.setCurrentIndex(workbooks_model.index(0, 0))

        logging.info('Loaded successfully {} files out of {}'.format(n_loaded_files, n_xlsx_files))

    def on_quit_application(self):
        """Event handler when the application is exited.
        """

        choice = QtWidgets.QMessageBox.question(self, 'Quit', "Do you really want to quit?", QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
        if choice == QtWidgets.QMessageBox.Yes:
            sys.exit()

    def on_update_property_plot(self, selected_property=None):
        """
        """

        selected_property = self._selected_property_combo.currentText()
        if not selected_property:
            return

        selected_indexes = self._animals_list.selectionModel().selectedIndexes()
        if not selected_indexes:
            return

        workbooks_model = self._animals_list.model()

        selected_workbooks = [workbooks_model.data(index, AnimalsDataModel.Workbook) for index in selected_indexes]

        selected_data = []
        for wb in selected_workbooks:
            property_slice = wb.get_property_slice(selected_property)
            selected_data.append((wb.basename, property_slice))

        selected_data = (selected_property, selected_data)
        self.update_property_plot.emit(selected_data)

    def on_update_time_plot(self, selected_time=None):
        """
        """

        selected_property = self._selected_property_combo.currentText()
        if not selected_property:
            return

        selected_time = self._selected_time_combo.currentText()
        if not selected_time:
            return

        selected_indexes = self._animals_list.selectionModel().selectedIndexes()
        if not selected_indexes:
            return

        workbooks_model = self._animals_list.model()

        selected_workbooks = [workbooks_model.data(index, AnimalsDataModel.Workbook) for index in selected_indexes]

        selected_data = []
        for wb in selected_workbooks:
            time_slice = wb.get_time_slice(selected_time)
            if selected_property not in time_slice.columns:
                value = np.nan
            else:
                value = time_slice[selected_property].loc[selected_time]
            selected_data.append((wb.basename, value))
        selected_data = (selected_property, selected_time, selected_data)

        self.update_time_plot.emit(selected_data)

    def on_select_animal(self, index):
        """Event fired when an animal is selected.

        Args:
            index (PyQt5.QtCore.QModelIndex): the index of the animal in the corresponding list view
        """

        selected_indexes = index.indexes()

        if not selected_indexes:
            return

        dataframe = self._animals_list.model().data(selected_indexes[0], AnimalsDataModel.Workbook).data

        workbook_data_model = WorkbookDataModel(dataframe)

        self._data_table.setModel(workbook_data_model)

        self.on_update_property_plot()

        self.on_update_time_plot()

    @ property
    def selected_property(self):

        return self._selected_property_combo.currentText()
